import openpyxl
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

sender_email="Pra0000kri@gmail.com"
password="zjql zjom hmam amzq"

file_path="BirthdayTracker.xlsx"
workbook=openpyxl.load_workbook(file_path)
sheet=workbook.active

# for row in sheet.iter_rows(min_row=2,values_only=True):
#     x=row[1]
#     x=x.month
#     print(x)
name=[]
email=[]
for row in sheet.iter_rows(min_row=2,values_only=True):
    x=row[1]
    if (x.month)==1:
        email.append(row[2])
        name.append(row[0])

for i in range(len(name)):
    message=MIMEMultipart()
    message["From"]=sender_email
    message["To"]=email[i]
    message["Subject"]="Happy Birth Month"

    body=f"Happy Birth Month to you {name[i]}.I hope you have a great year ahead."
    message.attach(MIMEText(body,"plain"))

    try:
        with smtplib.SMTP("smtp.gmail.com",587) as server:
            server.starttls()
            server.login(sender_email, password)
            server.sendmail(sender_email, email[i], message.as_string())
        print("Email sent successfully")
    except Exception as e:
        print(f"Error: {e}")
